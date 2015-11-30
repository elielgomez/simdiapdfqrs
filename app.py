import os
import sys
from PyQt4.QtGui import *
from PyQt4.QtCore import pyqtSlot
from PyQt4 import QtWebKit
import openpyxl as xls
from fpdf import FPDF
import pyqrcode
import string

class Window:
	a = QApplication(sys.argv)       
	# The QWidget widget is the base class of all user interface objects in PyQt4.
	w = QWidget()

	textbox = QLineEdit(w)
	textboxFileName = QLineEdit(w)
	progress = QProgressBar(w)
	label = QLabel(w)
    

	# Initialize
	def __init__(self,width,height,title):	
		# Set window size. 
		self.w.resize(width, height)
		# Set window title  
		self.w.setWindowTitle(title)

		self.textbox.move(100, 10)
		self.textbox.resize(280,20)

		self.textboxFileName.move(100,50)
		self.textboxFileName.resize(280,20)
		self.textboxFileName.setPlaceholderText("File Name...")

		self.label.move(10,105)
		self.label.resize(200,15)

		self.progress.setGeometry(10, 80, 405, 20)

	# On Click
	@pyqtSlot()
	def router_clicked(self,route):
		if route == 'exit':
			exit()
		elif route == 'msgbox_salir':
			result = self.messageBox_yesno('Salir','Desea salir?')
			if result == QMessageBox.Yes:
				exit()

		elif route == 'select_file':
			filename = self.openFileDialog()
			self.textbox.setText(filename)

		elif route == 'load_file':
			self.load_file(self.textbox.text())

	# On Press
	@pyqtSlot()
	def on_press(self):
		print('pressed')

	# On Released
	@pyqtSlot()
	def released(self):
		print('released')		

	# Add Button 
	def addButton(self,title,tooltip,x,y,route):
		btn = QPushButton(title, self.w) 
		btn.setToolTip(tooltip)
		# Send a text "route" to method by a lmabda function
		slotLambda = lambda: self.router_clicked(route)
		btn.clicked.connect(slotLambda)
		btn.resize(btn.sizeHint())
		btn.move(x,y)    

	#Show MessageBox Yes or No
	def messageBox_yesno(self, title, message):
		result = QMessageBox.question(self.w, title, message, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
		return result

	def messageBox(self,title,message):
		result = QMessageBox.information(self.w, title, message, QMessageBox.Ok)
		return result

	# Show Window 
	def show(self):
		self.w.show()
		sys.exit(self.a.exec_())

	def openFileDialog(self):
		filename = QFileDialog.getOpenFileName(self.w, 'Select File', '')
		return filename

	def load_file(self,file_name):
		if not os.path.exists('qrs'):
			os.makedirs('qrs')
		if self.textboxFileName.text() == '':
			self.messageBox('Filename', 'Type a filename')
		else:
			try:
				qrs = list()
				wb = xls.load_workbook(file_name)
				sheet = wb.worksheets[0]
				for row in range(2, sheet.max_row + 1):
					text_qr  = str(sheet['B' + str(row)].value) + '&' + str(sheet['A' + str(row)].value) + '&' + str(sheet['D' + str(row)].value) + '&' + str(sheet['C' + str(row)].value) + '&' + str(sheet['E' + str(row)].value) + '&' + str(sheet['F' + str(row)].value) + '&' + str(sheet['G' + str(row)].value) + '&' + str(sheet['H' + str(row)].value)
					trampa = str(sheet['D'+ str(row)].value)
					qrs.append(trampa)
					self.generate_qr(text_qr,trampa)
					self.progress.setValue((row*100)/sheet.max_row)
					self.label.setText('Reading Excel File...')
				self.generate_pdf(qrs,self.textboxFileName.text())

			except Exception as e:
				self.messageBox('Error', str(e))
				self.textbox.setText('')

	def generate_qr(self,data,filename):
		big_code = pyqrcode.create(data, error='H', version=None, mode=None, encoding='utf-8')
		big_code.png('qrs/'+filename+'.png', scale=5)

	def generate_pdf(self,qrs,pdf_filename):
		self.progress.setValue(0)
		self.label.setText('Writing PDF...')
		pdf=FPDF()
		pdf.add_page()
		col = 10
		row = 20
		pdf.set_font('Arial','B',20)
		pdf.text(10, 10, 'QRS')
		pdf.set_font('Arial','',5)
		c = 1
		for qr in qrs:
			if col > 250:
				row = row + 40
				col = 10
			pdf.text(col+3,row-3,qr)
			pdf.image('qrs/'+qr+'.png',col,row,30)
			col = col + 40
			self.progress.setValue((c*100)/len(qrs))
			c = c + 1
		try: 
			if not os.path.exists('pdfs'):
				os.makedirs('pdfs')
			pdf.output('pdfs/'+pdf_filename+'.pdf','F')
			self.messageBox('PDF', 'Pdf is Done! ' + 'pdfs/'+pdf_filename+'.pdf')
			self.label.setText('')

			path = os.getcwd()
			os.startfile(path.replace('\\','/')+'/pdfs/'+pdf_filename+'.pdf')

		except Exception as e:
			self.messageBox('PDF.', str(e))


def main():
	window = Window(420,120,"App")
	window.addButton("Select File", "Select File to Open",10,9,"select_file")
	window.addButton("Generate PDF", "Generates an PDF Document",9,50,"load_file")
	window.show()

if __name__ == '__main__':
	main()


